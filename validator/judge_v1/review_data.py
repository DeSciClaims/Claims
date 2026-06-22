from __future__ import annotations

import csv
import re
import xml.etree.ElementTree as ET
from zipfile import ZipFile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, Sequence


REVIEW_DECISIONS = ("accept", "revise", "reject")

NOTE_TAG_PATTERNS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("reference", (r"\breference\b", r"\bquote from a reference\b")),
    ("proposition", (r"\bproposition\b", r"\bnot (?:really )?a meaningful claim\b", r"\bpreview of the paper's main result\b")),
    ("missing_context", (r"\bcontext missing\b", r"\bimportant context missing\b", r"\blooking at only one isolated sentence\b")),
    ("missing_qualifier", (r"\bqualifier\b", r"\bqualifiers\b")),
    ("missing_detail", (r"\bdetail\b", r"\bdetailed information\b", r"\beffect sizes\b")),
    ("missing_condition", (r"\bcondition\b", r"\bcomparison group\b", r"\bthreshold\b")),
    ("missing_object", (r"\bobject was missing\b",)),
    ("bad_predicate", (r"\bpredicate\b",)),
    ("misplaced_argument", (r"\bsubject\(s\) is part of the object\b", r"\bpredicate containes subjects\b")),
    ("needs_split", (r"\btwo separate claims\b", r"\bseveral separate claims\b", r"\bindependently reported\b", r"\bu-shape\b")),
    ("not_result", (r"\bdescription of the dataset\b", r"\bstatistical model\b", r"\bintroduction for readers\b")),
    ("factually_truncated", (r"\bthrowing away\b", r"\bshortened form\b")),
)

NOTE_TARGETS = (
    "claim_text_type",
    "context_atomicity",
    "spo_fields",
    "claim_text_fidelity",
    "mixed_or_other",
)

NOTE_TARGET_PATTERNS: tuple[tuple[str, tuple[str, ...]], ...] = (
    (
        "claim_text_type",
        (
            r"\breference\b",
            r"\breferences\b",
            r"\bproposition\b",
            r"\bdefinition,? not a claim\b",
            r"\bthis is a definition, not a claim\b",
            r"\bnot a scientific claim\b",
            r"\bnot a meaningful claim\b",
            r"\bnot really a meaningful claim\b",
            r"\bdataset description\b",
            r"\bdescription of the dataset\b",
            r"\bmethod detail\b",
            r"\bstatistical model\b",
            r"\bheader of figure\b",
            r"\bquote from a reference\b",
            r"\bpreview of the paper's main result\b",
            r"\bdiscussion\b",
        ),
    ),
    (
        "context_atomicity",
        (
            r"\bcontext missing\b",
            r"\bimportant context missing\b",
            r"\bmeaningless without context\b",
            r"\ball relevant context\b",
            r"\ball meaningful context\b",
            r"\bimportant qualifier\b",
            r"\bqualifier\b",
            r"\bqualifiers\b",
            r"\bcondition was missing\b",
            r"\bcomparison group\b",
            r"\bthreshold\b",
            r"\btwo separate claims\b",
            r"\bseveral separate claims\b",
            r"\bindependently reported\b",
            r"\bu-shape\b",
            r"\bcan only be stated correctly by observing the context\b",
        ),
    ),
    (
        "spo_fields",
        (
            r"\bobject was missing\b",
            r"\bsubject was missing\b",
            r"\bpredicate\b",
            r"\bcomparison group was missing from the statement\b",
            r"\bsubjects?\) is part of the object\b",
            r"\bpredicate containes subjects\b",
        ),
    ),
    (
        "claim_text_fidelity",
        (
            r"\bthrowing away a lot of detailed information\b",
            r"\bshortened form is throwing away\b",
            r"\bmisunderstood the meaning\b",
            r"\bdid not understand the mathematical notation\b",
            r"\bmisses the statistical point\b",
            r"\bphrased as an empirical results\b",
        ),
    ),
)


@dataclass(frozen=True)
class ReviewedClaimRow:
    source_file: str
    paper_id: str
    claim_id: str
    section_title: str
    source_quote: str
    extracted_claim_text: str
    extracted_subject: str
    extracted_predicate: str
    extracted_object: str
    corrected_claim_text: str
    corrected_subject: str
    corrected_predicate: str
    corrected_object: str
    reviewer_decision: str
    reviewer_notes: str
    reviewer_note_tags: tuple[str, ...] = field(default_factory=tuple)

    @property
    def gold_claim_text(self) -> str:
        return self.corrected_claim_text or self.extracted_claim_text

    @property
    def gold_subject(self) -> str:
        return self.corrected_subject or self.extracted_subject

    @property
    def gold_predicate(self) -> str:
        return self.corrected_predicate or self.extracted_predicate

    @property
    def gold_object(self) -> str:
        return self.corrected_object or self.extracted_object


@dataclass(frozen=True)
class ReviewedQuoteGroup:
    group_id: str
    paper_id: str
    section_title: str
    source_quote: str
    rows: tuple[ReviewedClaimRow, ...]

    @property
    def note_tags(self) -> tuple[str, ...]:
        return tuple(_unique_preserve(tag for row in self.rows for tag in row.reviewer_note_tags))

    @property
    def note_texts(self) -> tuple[str, ...]:
        return tuple(_unique_preserve(row.reviewer_notes for row in self.rows if row.reviewer_notes))

    @property
    def target_claim_texts(self) -> tuple[str, ...]:
        gold_claims = [
            row.gold_claim_text
            for row in self.rows
            if row.reviewer_decision in {"accept", "revise"} and row.gold_claim_text
        ]
        if gold_claims:
            return tuple(_unique_preserve(gold_claims))
        fallback_claims = [row.gold_claim_text for row in self.rows if row.gold_claim_text]
        return tuple(_unique_preserve(fallback_claims))

    @property
    def target_claim_count(self) -> int:
        split_floor = 2 if "needs_split" in self.note_tags else 1
        return max(len(self.target_claim_texts), split_floor)


def default_reviewed_files_dir(repo_root: Path) -> Path:
    return repo_root / "examples" / "reviewed_files"


def normalize_decision(value: str) -> str:
    decision = _normalize_whitespace(value).lower()
    if decision not in REVIEW_DECISIONS:
        raise ValueError(f"Unexpected reviewer decision: {value!r}")
    return decision


def normalize_optional_decision(value: str) -> str:
    decision = _normalize_whitespace(value).lower()
    if not decision:
        return ""
    if decision not in REVIEW_DECISIONS:
        raise ValueError(f"Unexpected reviewer decision: {value!r}")
    return decision


def infer_note_tags(note: str) -> tuple[str, ...]:
    normalized_note = _normalize_whitespace(note).lower()
    if not normalized_note:
        return ()

    tags: list[str] = []
    for tag, patterns in NOTE_TAG_PATTERNS:
        if any(re.search(pattern, normalized_note) for pattern in patterns):
            tags.append(tag)
    return tuple(tags)


def infer_note_target(note: str) -> str:
    normalized_note = _normalize_whitespace(note).lower()
    if not normalized_note:
        return "mixed_or_other"

    matched_targets: list[str] = []
    for target, patterns in NOTE_TARGET_PATTERNS:
        if any(re.search(pattern, normalized_note) for pattern in patterns):
            matched_targets.append(target)

    if not matched_targets:
        return "mixed_or_other"

    # Prefer the most actionable buckets for judge design.
    for priority_target in ("context_atomicity", "claim_text_type", "spo_fields", "claim_text_fidelity"):
        if priority_target in matched_targets:
            return priority_target
    return "mixed_or_other"


def infer_section_type(section_title: str) -> str:
    normalized = _normalize_whitespace(section_title).lower()
    if any(token in normalized for token in ("abstract", "introduction", "background")):
        return "INTRO"
    if any(token in normalized for token in ("method", "materials", "data", "sample", "measure")):
        return "METHODS"
    if any(token in normalized for token in ("result", "finding")):
        return "RESULTS"
    if any(token in normalized for token in ("discussion", "conclusion", "implication")):
        return "DISCUSSION"
    if "table" in normalized:
        return "TABLE"
    return "OTHER"


def load_reviewed_claim_rows(reviewed_files_dir: Path) -> list[ReviewedClaimRow]:
    paths = sorted([*reviewed_files_dir.glob("*.csv"), *reviewed_files_dir.glob("*.xlsx")])
    return load_reviewed_claim_rows_from_paths(paths)


def load_reviewed_claim_rows_from_file(reviewed_file: Path) -> list[ReviewedClaimRow]:
    return load_reviewed_claim_rows_from_paths([reviewed_file])


def load_reviewed_claim_rows_from_paths(paths: Sequence[Path]) -> list[ReviewedClaimRow]:
    rows: list[ReviewedClaimRow] = []
    for source_path in paths:
        for raw_row in read_reviewer_rows(source_path):
            reviewer_notes = _normalize_whitespace(raw_row.get("reviewer_notes", ""))
            rows.append(
                ReviewedClaimRow(
                    source_file=source_path.name,
                    paper_id=_normalize_whitespace(raw_row.get("paper_id", "")),
                    claim_id=_normalize_whitespace(raw_row.get("claim_id", "")),
                    section_title=_normalize_whitespace(raw_row.get("section_title", "")),
                    source_quote=_normalize_whitespace(raw_row.get("source_quote", "")),
                    extracted_claim_text=_normalize_whitespace(raw_row.get("extracted_claim_text", "")),
                    extracted_subject=_normalize_whitespace(raw_row.get("extracted_subject", "")),
                    extracted_predicate=_normalize_whitespace(raw_row.get("extracted_predicate", "")),
                    extracted_object=_normalize_whitespace(raw_row.get("extracted_object", "")),
                    corrected_claim_text=_normalize_whitespace(raw_row.get("corrected_claim_text", "")),
                    corrected_subject=_normalize_whitespace(raw_row.get("corrected_subject", "")),
                    corrected_predicate=_normalize_whitespace(raw_row.get("corrected_predicate", "")),
                    corrected_object=_normalize_whitespace(raw_row.get("corrected_object", "")),
                    reviewer_decision=normalize_optional_decision(raw_row.get("reviewer_decision", "")),
                    reviewer_notes=reviewer_notes,
                    reviewer_note_tags=infer_note_tags(reviewer_notes),
                )
            )
    return rows


def read_reviewer_rows(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as handle:
            return [_stringify_row(row) for row in csv.DictReader(handle)]
    if suffix == ".xlsx":
        return _read_xlsx_rows(path)
    return []


def group_rows_by_quote(rows: Sequence[ReviewedClaimRow]) -> list[ReviewedQuoteGroup]:
    grouped: dict[tuple[str, str, str], list[ReviewedClaimRow]] = defaultdict(list)
    for row in rows:
        key = (row.paper_id, row.section_title, row.source_quote)
        grouped[key].append(row)

    quote_groups: list[ReviewedQuoteGroup] = []
    for index, ((paper_id, section_title, source_quote), group_rows) in enumerate(sorted(grouped.items()), start=1):
        quote_groups.append(
            ReviewedQuoteGroup(
                group_id=f"group_{index:03d}",
                paper_id=paper_id,
                section_title=section_title,
                source_quote=source_quote,
                rows=tuple(group_rows),
            )
        )
    return quote_groups


def split_rows_train_val_test(rows: Sequence[ReviewedClaimRow]) -> tuple[list[ReviewedClaimRow], list[ReviewedClaimRow], list[ReviewedClaimRow]]:
    groups = group_rows_by_quote(rows)
    if not groups:
        return [], [], []

    total = len(groups)
    train_cutoff = max(1, round(total * 0.6))
    val_cutoff = max(train_cutoff + 1, round(total * 0.8)) if total > 2 else min(total, train_cutoff + 1)

    train_keys = {group.group_id for group in groups[:train_cutoff]}
    val_keys = {group.group_id for group in groups[train_cutoff:val_cutoff]}

    row_to_group_id = {
        id(row): group.group_id
        for group in groups
        for row in group.rows
    }
    train_rows = [row for row in rows if row_to_group_id.get(id(row)) in train_keys]
    val_rows = [row for row in rows if row_to_group_id.get(id(row)) in val_keys]
    test_rows = [row for row in rows if row_to_group_id.get(id(row)) not in train_keys | val_keys]
    return train_rows, val_rows, test_rows


def split_quote_groups_train_val_test(
    groups: Sequence[ReviewedQuoteGroup],
) -> tuple[list[ReviewedQuoteGroup], list[ReviewedQuoteGroup], list[ReviewedQuoteGroup]]:
    total = len(groups)
    if total == 0:
        return [], [], []
    train_cutoff = max(1, round(total * 0.6))
    val_cutoff = max(train_cutoff + 1, round(total * 0.8)) if total > 2 else min(total, train_cutoff + 1)
    return list(groups[:train_cutoff]), list(groups[train_cutoff:val_cutoff]), list(groups[val_cutoff:])


def summarize_reviewer_notes(rows: Sequence[ReviewedClaimRow]) -> dict[str, object]:
    decision_counts = Counter(row.reviewer_decision for row in rows)
    tag_counts = Counter(tag for row in rows for tag in row.reviewer_note_tags)
    raw_note_counts = Counter(row.reviewer_notes for row in rows if row.reviewer_notes)
    note_target_counts = Counter(infer_note_target(row.reviewer_notes) for row in rows if row.reviewer_notes)
    return {
        "total_rows": len(rows),
        "decision_counts": dict(decision_counts),
        "tag_counts": dict(tag_counts),
        "note_target_counts": dict(note_target_counts),
        "top_reviewer_notes": raw_note_counts.most_common(25),
    }


def render_reviewer_note_report(rows: Sequence[ReviewedClaimRow]) -> str:
    summary = summarize_reviewer_notes(rows)
    lines = [
        f"Total reviewed rows: {summary['total_rows']}",
        f"Decision counts: {summary['decision_counts']}",
        "Tag counts:",
    ]
    for tag, count in sorted(summary["tag_counts"].items()):
        lines.append(f"- {tag}: {count}")
    lines.append("Top reviewer notes:")
    for note, count in summary["top_reviewer_notes"]:
        lines.append(f"- {count}x {note}")
    return "\n".join(lines)


def _read_xlsx_rows(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return _read_xlsx_rows_fallback(path)

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            return []
        fieldnames = [str(value or "").strip() for value in header]
        rows: list[dict[str, str]] = []
        for values in rows_iter:
            row = {
                field: "" if value is None else str(value)
                for field, value in zip(fieldnames, values, strict=False)
                if field
            }
            if any(str(value).strip() for value in row.values()):
                rows.append(_stringify_row(row))
        return rows
    finally:
        workbook.close()


def _read_xlsx_rows_fallback(path: Path) -> list[dict[str, str]]:
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }

    def _sheet_target(target: str) -> str:
        if target.startswith("/"):
            return target.lstrip("/")
        if target.startswith("xl/"):
            return target
        return f"xl/{target}"

    def _column_letters(cell_ref: str) -> str:
        letters: list[str] = []
        for char in cell_ref:
            if char.isalpha():
                letters.append(char)
            else:
                break
        return "".join(letters)

    with ZipFile(path) as archive:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in shared_root.findall("main:si", ns):
                shared_strings.append("".join(node.text or "" for node in item.iterfind(".//main:t", ns)))

        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        relationships = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in relationships_root.findall("rel:Relationship", ns)
        }
        first_sheet = workbook_root.find("main:sheets/main:sheet", ns)
        if first_sheet is None:
            return []
        relationship_id = first_sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
        sheet_root = ET.fromstring(archive.read(_sheet_target(relationships[relationship_id])))

        parsed_rows: list[dict[str, str]] = []
        for row in sheet_root.findall(".//main:sheetData/main:row", ns):
            parsed: dict[str, str] = {}
            for cell in row.findall("main:c", ns):
                column = _column_letters(cell.attrib.get("r", ""))
                if not column:
                    continue
                cell_type = cell.attrib.get("t", "")
                value_node = cell.find("main:v", ns)
                inline_node = cell.find("main:is", ns)
                if cell_type == "s" and value_node is not None and value_node.text is not None:
                    value = shared_strings[int(value_node.text)]
                elif cell_type == "inlineStr" and inline_node is not None:
                    value = "".join(node.text or "" for node in inline_node.iterfind(".//main:t", ns))
                elif value_node is not None and value_node.text is not None:
                    value = value_node.text
                else:
                    value = ""
                parsed[column] = value
            if parsed:
                parsed_rows.append(parsed)

    if not parsed_rows:
        return []

    header_row = parsed_rows[0]
    header_columns = sorted(header_row.keys(), key=_spreadsheet_column_key)
    fieldnames = [str(header_row.get(column, "")).strip() for column in header_columns]
    rows: list[dict[str, str]] = []
    for parsed_row in parsed_rows[1:]:
        row = {
            field: str(parsed_row.get(column, "") or "")
            for field, column in zip(fieldnames, header_columns, strict=False)
            if field
        }
        if any(value.strip() for value in row.values()):
            rows.append(_stringify_row(row))
    return rows


def _spreadsheet_column_key(column: str) -> tuple[int, str]:
    return (len(column), column)


def _stringify_row(row: dict[str, Any]) -> dict[str, str]:
    return {
        str(key): "" if value is None else str(value)
        for key, value in row.items()
    }


def _normalize_whitespace(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def _unique_preserve(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    unique_values: list[str] = []
    for value in values:
        normalized = _normalize_whitespace(value)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        unique_values.append(normalized)
    return unique_values
