from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from .profiles import CLAIM_PROFILE, CLAIM_PROFILES, EVIDENCE_METHOD_PROFILES, FIELD_POLICIES, normalize_claim_profile


COLUMN_WIDTHS = {
    "paper_id": 18,
    "claim_id": 24,
    "claim_profile": 28,
    "section_title": 28,
    "source_quote": 100,
    "extracted_claim_text": 60,
    "extracted_subject": 28,
    "extracted_predicate": 24,
    "extracted_object": 42,
    "prototype_claim_json": 70,
    "extracted_context_summary": 42,
    "extracted_context_json": 52,
    "extracted_details_summary": 48,
    "extracted_details_json": 60,
    "extractor_metadata_summary": 52,
    "extractor_metadata_json": 64,
    "linked_evidence_ids": 28,
    "group_evidence_count": 16,
    "group_evidence_summary": 48,
    "group_evidence_items_json": 60,
    "group_link_count": 16,
    "group_links_json": 60,
    "corrected_claim_text": 60,
    "corrected_subject": 28,
    "corrected_predicate": 24,
    "corrected_object": 42,
    "reviewer_decision": 18,
    "reviewer_notes": 50,
}

CLAIM_CONTEXT_KEY_ALIASES = {
    "comparison_group": "comparator",
    "comparison details": "comparator",
    "cohort_info": "cohort",
}

CLAIM_CONTEXT_KEYS = set(CLAIM_PROFILE.get("allowed_context_keys", []))
CLAIM_DETAIL_KEYS = set(CLAIM_PROFILE.get("allowed_details_keys", []))
EVIDENCE_CONTEXT_KEYS = {
    key
    for profile in EVIDENCE_METHOD_PROFILES.values()
    for key in profile.get("allowed_context_keys", [])
}
EVIDENCE_DETAIL_KEYS = {
    key
    for profile in EVIDENCE_METHOD_PROFILES.values()
    for key in profile.get("allowed_details_keys", [])
}
EXTRACTOR_METADATA_KEYS = {
    "graph_subject",
    "graph_predicate",
    "graph_object",
    "surface_subject",
    "surface_predicate",
    "surface_object",
    "claim_candidate_score",
    "spo_candidate_score",
    "joint_score",
    "source_group_evidence_count",
    "source_group_link_count",
    "claim_candidate_rationale",
    "claim_candidate_context_notes",
    "spo_candidate_rationale",
    "qualifier_notes",
    "repair_reason",
    "repair_action",
    "original_claim_text",
    "original_subject",
    "original_predicate",
    "original_object",
}


def write_reviewer_rows(
    rows: list[dict[str, Any]],
    output_path: Path,
    fieldnames: list[str],
    *,
    xlsx: bool,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if xlsx:
        write_xlsx(rows, output_path, fieldnames)
    else:
        write_csv(rows, output_path, fieldnames)


def write_csv(rows: list[dict[str, Any]], output_path: Path, fieldnames: list[str]) -> None:
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, Any]], output_path: Path, fieldnames: list[str]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - depends on env
        raise RuntimeError(
            "openpyxl is required for --xlsx export. Install it with `pip install openpyxl`."
        ) from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reviewer Claims"

    worksheet.append(fieldnames)
    for row in rows:
        worksheet.append([row.get(field, "") for field in fieldnames])

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_alignment

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for index, field in enumerate(fieldnames, start=1):
        letter = get_column_letter(index)
        worksheet.column_dimensions[letter].width = COLUMN_WIDTHS.get(field, 22)

    has_source_quote = "source_quote" in fieldnames
    for row_index in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = 90 if has_source_quote else 45

    workbook.save(output_path)


def serialize_export_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    try:
        return json.dumps(_jsonable(value), ensure_ascii=False, sort_keys=True)
    except Exception:
        return str(value)


def _jsonable(value: Any) -> Any:
    if hasattr(value, "model_dump"):
        return value.model_dump(mode="json")
    if isinstance(value, dict):
        return {str(key): _jsonable(raw) for key, raw in value.items()}
    if isinstance(value, (list, tuple)):
        return [_jsonable(raw) for raw in value]
    return value


def linked_evidence_ids_for_claim(claim_id: str, links: list[dict[str, Any]]) -> str:
    evidence_ids = [
        str(link.get("evidence_id", "")).strip()
        for link in links
        if str(link.get("claim_id", "")).strip() == str(claim_id).strip()
        and str(link.get("evidence_id", "")).strip()
    ]
    return "; ".join(dict.fromkeys(evidence_ids))


def summarize_context(context: Any) -> str:
    if not isinstance(context, dict) or not context:
        return ""
    parts: list[str] = []
    for key, value in context.items():
        if hasattr(value, "model_dump"):
            text = str(value.model_dump(mode="json").get("value", "")).strip()
        elif isinstance(value, dict):
            text = str(value.get("value", "")).strip()
        else:
            text = str(value).strip()
        if text:
            parts.append(f"{key}={text}")
    return "; ".join(parts)


def summarize_details(details: Any, *, exclude_keys: set[str] | None = None) -> str:
    if not isinstance(details, dict) or not details:
        return ""
    exclude = exclude_keys or set()
    parts: list[str] = []
    for key, value in details.items():
        if key in exclude:
            continue
        if isinstance(value, (dict, list)):
            text = serialize_export_value(value)
        else:
            text = str(value).strip()
        if text:
            parts.append(f"{key}={text}")
    return "; ".join(parts[:8])


def summarize_evidence_items(evidence_items: list[dict[str, Any]]) -> str:
    if not evidence_items:
        return ""
    parts: list[str] = []
    for item in evidence_items[:3]:
        summary = str(item.get("summary_text", "")).strip()
        method = ""
        evidence_method = item.get("evidence_method")
        if isinstance(evidence_method, dict):
            method = str(evidence_method.get("value", "")).strip()
        if summary and method:
            parts.append(f"{method}: {summary}")
        elif summary:
            parts.append(summary)
        elif method:
            parts.append(method)
    if len(evidence_items) > 3:
        parts.append(f"... (+{len(evidence_items) - 3} more)")
    return " | ".join(parts)


NUMERIC_PAYLOAD_KEYS = {
    "effect_size",
    "effect_direction",
    "statistical_significance",
    "count",
    "sample_size",
    "study_count",
    "p_value",
    "ci_low",
    "ci_high",
    "variance_explained",
    "lag",
    "correlation_value",
    "value",
}


def _short_concept_phrase(text: str) -> str:
    cleaned = " ".join(str(text or "").replace("\n", " ").split()).strip(" .;:")
    if len(cleaned.split()) > 10:
        return ""
    if any(token in cleaned for token in ("=", "<", ">", "%")):
        return ""
    if cleaned.replace(".", "", 1).isdigit():
        return ""
    return cleaned


def _field_role(path: str, *, claim_profile: str | None = None, evidence_method: str | None = None) -> str:
    if claim_profile:
        override = FIELD_POLICIES.get("profile_field_role_overrides", {}).get(claim_profile, {}).get(path)
        if override:
            return str(override)
    if evidence_method:
        override = FIELD_POLICIES.get("evidence_method_field_role_overrides", {}).get(evidence_method, {}).get(path)
        if override:
            return str(override)
    return str(FIELD_POLICIES.get("claim_field_roles", {}).get(path) or FIELD_POLICIES.get("evidence_field_roles", {}).get(path) or "")


def normalize_semantic_field(
    value: Any,
    *,
    default_entity_type: str = "entity",
    field_path: str | None = None,
    claim_profile: str | None = None,
    evidence_method: str | None = None,
) -> dict[str, Any]:
    if isinstance(value, dict):
        raw_text = str(value.get("value", "")).strip()
        role = _field_role(field_path or "", claim_profile=claim_profile, evidence_method=evidence_method)
        text = _short_concept_phrase(raw_text) if role == "ontology-target" else raw_text
        return {
            "value": text,
            "entity_type": str(value.get("entity_type", "")).strip() or default_entity_type,
            "ontology": value.get("ontology"),
        }
    raw_text = str(value or "").strip()
    role = _field_role(field_path or "", claim_profile=claim_profile, evidence_method=evidence_method)
    return {
        "value": _short_concept_phrase(raw_text) if role == "ontology-target" else raw_text,
        "entity_type": default_entity_type,
        "ontology": None,
    }


def normalize_context_payload(context: Any, *, claim_profile: str | None = None) -> dict[str, Any]:
    normalized_context, _, _ = normalize_claim_payload_parts(context=context, details=None, claim_profile=claim_profile)
    return normalized_context


def normalize_details_payload(
    details: Any,
    *,
    claim_profile: str | None = None,
    subject: Any | None = None,
    predicate: Any | None = None,
    object_: Any | None = None,
) -> dict[str, Any]:
    _, normalized_details, _ = normalize_claim_payload_parts(
        context=None,
        details=details,
        claim_profile=claim_profile,
        subject=subject,
        predicate=predicate,
        object_=object_,
    )
    return normalized_details


def normalize_extractor_metadata_payload(
    details: Any,
    *,
    claim_profile: str | None = None,
    subject: Any | None = None,
    predicate: Any | None = None,
    object_: Any | None = None,
) -> dict[str, Any]:
    _, _, metadata = normalize_claim_payload_parts(
        context=None,
        details=details,
        claim_profile=claim_profile,
        subject=subject,
        predicate=predicate,
        object_=object_,
    )
    return metadata


def normalize_evidence_context_payload(
    context: Any,
    *,
    details: Any | None = None,
    evidence_method: Any | None = None,
) -> dict[str, Any]:
    normalized_context, _ = normalize_evidence_payload_parts(
        context=context,
        details=details,
        evidence_method=evidence_method,
    )
    return normalized_context


def normalize_evidence_details_payload(
    details: Any,
    *,
    context: Any | None = None,
    evidence_method: Any | None = None,
) -> dict[str, Any]:
    _, normalized_details = normalize_evidence_payload_parts(
        context=context,
        details=details,
        evidence_method=evidence_method,
    )
    return normalized_details


def normalize_claim_payload_parts(
    *,
    context: Any,
    details: Any,
    claim_profile: str | None = None,
    subject: Any | None = None,
    predicate: Any | None = None,
    object_: Any | None = None,
) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    raw_context = context if isinstance(context, dict) else {}
    raw_details = details if isinstance(details, dict) else {}
    profile_name = normalize_claim_profile(claim_profile)
    profile = CLAIM_PROFILES.get(profile_name, CLAIM_PROFILES["generic_result"])
    allowed_context_keys = set(profile.get("allowed_context_keys", [])) or CLAIM_CONTEXT_KEYS
    allowed_detail_keys = set(profile.get("allowed_details_keys", [])) or CLAIM_DETAIL_KEYS
    forbidden_context_keys = set(profile.get("forbidden_context_keys", []))
    forbidden_detail_keys = set(profile.get("forbidden_details_keys", []))

    normalized_context: dict[str, Any] = {}
    normalized_details: dict[str, Any] = {}
    extractor_metadata: dict[str, Any] = {}

    for raw_key, raw_value in raw_context.items():
        key = _canonicalize_claim_key(raw_key, raw_value)
        if not key:
            continue
        if key in allowed_detail_keys and key not in forbidden_detail_keys:
            normalized_details[key] = _normalize_detail_value(
                raw_value,
                field_path=f"claim.details.{key}",
                claim_profile=profile_name,
            )
        elif key in forbidden_context_keys:
            extractor_metadata.setdefault("forbidden_context", {})[str(raw_key)] = raw_value
        elif key in allowed_context_keys:
            normalized_context[key] = normalize_semantic_field(
                raw_value,
                default_entity_type=key,
                field_path=f"claim.context.{key}",
                claim_profile=profile_name,
            )
        else:
            extractor_metadata.setdefault("legacy_context", {})[str(raw_key)] = raw_value

    graph_subject = raw_details.get("graph_subject")
    graph_predicate = raw_details.get("graph_predicate")
    graph_object = raw_details.get("graph_object")
    surface_subject = raw_details.get("surface_subject")
    surface_predicate = raw_details.get("surface_predicate")
    surface_object = raw_details.get("surface_object")

    if any([graph_subject, graph_predicate, graph_object]):
        extractor_metadata["graph_projection"] = {
            "subject": normalize_semantic_field(graph_subject or subject, default_entity_type="entity", field_path="claim.subject", claim_profile=profile_name),
            "predicate": normalize_semantic_field(graph_predicate or predicate, default_entity_type="predicate", field_path="claim.predicate", claim_profile=profile_name),
            "object": normalize_semantic_field(graph_object or object_, default_entity_type="entity", field_path="claim.object", claim_profile=profile_name),
        }

    if any([surface_subject, surface_predicate, surface_object]):
        extractor_metadata["surface_form"] = {
            "subject": normalize_semantic_field(surface_subject or subject, default_entity_type="entity", field_path="claim.subject", claim_profile=profile_name),
            "predicate": normalize_semantic_field(surface_predicate or predicate, default_entity_type="predicate", field_path="claim.predicate", claim_profile=profile_name),
            "object": normalize_semantic_field(surface_object or object_, default_entity_type="entity", field_path="claim.object", claim_profile=profile_name),
        }

    candidate_scoring_keys = {
        "claim_candidate_score",
        "spo_candidate_score",
        "joint_score",
        "source_group_evidence_count",
        "source_group_link_count",
    }
    candidate_scoring = {key: raw_details[key] for key in candidate_scoring_keys if key in raw_details}
    if candidate_scoring:
        extractor_metadata["candidate_selection"] = candidate_scoring

    note_keys = {
        "claim_candidate_rationale",
        "claim_candidate_context_notes",
        "spo_candidate_rationale",
        "qualifier_notes",
        "repair_reason",
        "repair_action",
        "original_claim_text",
        "original_subject",
        "original_predicate",
        "original_object",
    }
    notes = {key: raw_details[key] for key in note_keys if key in raw_details}
    if notes:
        extractor_metadata["generation_notes"] = notes

    used_keys = EXTRACTOR_METADATA_KEYS
    for raw_key, raw_value in raw_details.items():
        if raw_key in used_keys:
            continue
        key = _canonicalize_claim_key(raw_key, raw_value)
        if key in forbidden_detail_keys:
            extractor_metadata.setdefault("forbidden_details", {})[raw_key] = raw_value
        elif key in allowed_detail_keys:
            normalized_details[key] = _normalize_detail_value(
                raw_value,
                field_path=f"claim.details.{key}",
                claim_profile=profile_name,
            )
        elif key in allowed_context_keys and key not in normalized_context:
            normalized_context[key] = normalize_semantic_field(
                raw_value,
                default_entity_type=key,
                field_path=f"claim.context.{key}",
                claim_profile=profile_name,
            )
        else:
            extractor_metadata.setdefault("extra", {})[raw_key] = raw_value

    return normalized_context, normalized_details, extractor_metadata


def normalize_evidence_payload_parts(
    *,
    context: Any,
    details: Any,
    evidence_method: Any | None = None,
) -> tuple[dict[str, Any], dict[str, Any]]:
    raw_context = context if isinstance(context, dict) else {}
    raw_details = details if isinstance(details, dict) else {}

    evidence_method_value = ""
    if isinstance(evidence_method, dict):
        evidence_method_value = str(evidence_method.get("value", "")).strip()
    else:
        evidence_method_value = str(evidence_method or "").strip()
    method_profile = EVIDENCE_METHOD_PROFILES.get(evidence_method_value, {})

    allowed_context_keys = set(method_profile.get("allowed_context_keys", [])) or EVIDENCE_CONTEXT_KEYS
    allowed_detail_keys = set(method_profile.get("allowed_details_keys", [])) or EVIDENCE_DETAIL_KEYS

    normalized_context: dict[str, Any] = {}
    normalized_details: dict[str, Any] = {}

    for raw_key, raw_value in raw_context.items():
        key = _canonicalize_claim_key(raw_key, raw_value)
        if not key:
            continue
        if key in allowed_detail_keys:
            normalized_details[key] = _normalize_detail_value(
                raw_value,
                field_path=f"evidence.details.{key}",
                evidence_method=evidence_method_value,
            )
        elif key in allowed_context_keys:
            normalized_context[key] = normalize_semantic_field(
                raw_value,
                default_entity_type=key,
                field_path=f"evidence.context.{key}",
                evidence_method=evidence_method_value,
            )

    for raw_key, raw_value in raw_details.items():
        key = _canonicalize_claim_key(raw_key, raw_value)
        if not key:
            continue
        if key in allowed_context_keys and key not in normalized_context:
            normalized_context[key] = normalize_semantic_field(
                raw_value,
                default_entity_type=key,
                field_path=f"evidence.context.{key}",
                evidence_method=evidence_method_value,
            )
        elif key in allowed_detail_keys:
            normalized_details[key] = _normalize_detail_value(
                raw_value,
                field_path=f"evidence.details.{key}",
                evidence_method=evidence_method_value,
            )

    return normalized_context, normalized_details


def _normalize_detail_value(
    value: Any,
    *,
    field_path: str | None = None,
    claim_profile: str | None = None,
    evidence_method: str | None = None,
) -> Any:
    if not isinstance(value, dict):
        role = _field_role(field_path or "", claim_profile=claim_profile, evidence_method=evidence_method)
        if role == "ontology-target" and isinstance(value, str):
            return _short_concept_phrase(value)
        return value
    return {
        str(key): _normalize_detail_value(
            raw,
            field_path=f"{field_path}.{key}" if field_path else str(key),
            claim_profile=claim_profile,
            evidence_method=evidence_method,
        )
        for key, raw in value.items()
        if str(key).strip()
    }


def _canonicalize_claim_key(raw_key: Any, raw_value: Any) -> str:
    key = str(raw_key or "").strip()
    if not key:
        return ""
    canonical = CLAIM_CONTEXT_KEY_ALIASES.get(key, key)
    if canonical == "evidence_basis":
        text = ""
        if isinstance(raw_value, dict):
            text = str(raw_value.get("text") or raw_value.get("value") or "").lower()
        else:
            text = str(raw_value or "").lower()
        return "citation_context" if any(token in text for token in ("literature", "reference", "cited")) else "analysis_context"
    return canonical


def build_prototype_claim_payload(
    *,
    claim_id: str,
    paper_id: str,
    claim_text: str,
    subject: Any,
    predicate: Any,
    object_: Any,
    claim_kind: str = "",
    claim_profile: str | None = None,
    epistemic_status: str = "",
    support_origin: str = "",
    source_span_ids: list[str] | None = None,
    context: Any = None,
    details: Any = None,
    extractor_confidence: Any = None,
) -> dict[str, Any]:
    normalized_context, normalized_details, _ = normalize_claim_payload_parts(
        context=context,
        details=details,
        claim_profile=claim_profile,
        subject=subject,
        predicate=predicate,
        object_=object_,
    )
    return {
        "claim_id": claim_id,
        "paper_id": paper_id,
        "claim_text": claim_text,
        "subject": normalize_semantic_field(
            subject,
            default_entity_type="entity",
            field_path="claim.subject",
            claim_profile=claim_profile,
        ),
        "predicate": normalize_semantic_field(
            predicate,
            default_entity_type="predicate",
            field_path="claim.predicate",
            claim_profile=claim_profile,
        ),
        "object": normalize_semantic_field(
            object_,
            default_entity_type="entity",
            field_path="claim.object",
            claim_profile=claim_profile,
        ),
        "claim_kind": claim_kind,
        "claim_profile": normalize_claim_profile(claim_profile),
        "epistemic_status": epistemic_status,
        "support_origin": support_origin,
        "source_span_ids": list(source_span_ids or []),
        "context": normalized_context,
        "details": normalized_details,
        "extractor_confidence": extractor_confidence,
    }
